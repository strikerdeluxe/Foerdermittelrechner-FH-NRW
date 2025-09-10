import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

class FoerdermittelRechner:
    def __init__(self, gesamtsumme):
        """
        Initialisiert den Fördermittelrechner
        
        Args:
            gesamtsumme: Gesamte zu verteilende Fördersumme
        """
        self.gesamtsumme = gesamtsumme
        self.mindestbetrag = 12500
        self.sockelbetrag_prozent = 0.5
        self.kommunen_daten = []
        self.berechnungs_log = []
        
    def kommune_hinzufuegen(self, name, wert_2019, kinder_u3):
        """
        Fügt eine Kommune mit ihren Daten hinzu
        
        Args:
            name: Name der Kommune
            wert_2019: Förderwert aus 2019
            kinder_u3: Anzahl der Kinder U3 im SGB-II-Bezug
        """
        self.kommunen_daten.append({
            'Name': name,
            'Wert_2019': wert_2019,
            'Kinder_U3': kinder_u3,
            'Status': 'In Berechnung'
        })
    
    def berechne_verteilung(self):
        """
        Führt die iterative Berechnung der Fördermittelverteilung durch
        """
        # Erstelle DataFrame
        df = pd.DataFrame(self.kommunen_daten)
        df['Sockelbetrag'] = df['Wert_2019'] * self.sockelbetrag_prozent
        df['U3_Anteil'] = 0
        df['Zwischensumme'] = 0
        df['Endbetrag'] = 0
        df['Runde'] = 0
        df['Erste_Berechnung'] = 0
        
        # Tracking für Iterationen
        fixierte_kommunen = set()
        runde = 0
        max_runden = 10  # Sicherheitslimit
        
        while runde < max_runden:
            runde += 1
            print(f"\n=== Runde {runde} ===")
            
            # Filter aktive Kommunen (nicht fixiert)
            aktive_mask = ~df['Name'].isin(fixierte_kommunen)
            aktive_df = df[aktive_mask].copy()
            
            if len(aktive_df) == 0:
                print("Alle Kommunen wurden bearbeitet.")
                break
            
            # Berechne Sockelbeträge für aktive Kommunen
            summe_sockel = aktive_df['Sockelbetrag'].sum()
            
            # Berechne Restbudget
            bereits_fixiert_summe = df[df['Name'].isin(fixierte_kommunen)]['Endbetrag'].sum()
            verfuegbares_budget = self.gesamtsumme - bereits_fixiert_summe
            restbudget = verfuegbares_budget - summe_sockel
            
            print(f"Verfügbares Budget: {verfuegbares_budget:,.2f} €")
            print(f"Summe Sockelbeträge: {summe_sockel:,.2f} €")
            print(f"Restbudget für U3-Verteilung: {restbudget:,.2f} €")
            
            # Berechne U3-Multiplikator
            summe_u3 = aktive_df['Kinder_U3'].sum()
            if summe_u3 > 0:
                multiplikator = restbudget / summe_u3
            else:
                multiplikator = 0
            
            print(f"U3-Kinder gesamt: {summe_u3}")
            print(f"Multiplikator: {multiplikator:,.2f} €/Kind")
            
            # Berechne U3-Anteile und Zwischensummen
            for idx in aktive_df.index:
                df.loc[idx, 'U3_Anteil'] = df.loc[idx, 'Kinder_U3'] * multiplikator
                df.loc[idx, 'Zwischensumme'] = df.loc[idx, 'Sockelbetrag'] + df.loc[idx, 'U3_Anteil']
                df.loc[idx, 'Runde'] = runde
                
                # Speichere erste Berechnung
                if runde == 1:
                    df.loc[idx, 'Erste_Berechnung'] = df.loc[idx, 'Zwischensumme']
            
            # Prüfe Mindestbeträge
            neue_fixierungen = []
            for idx in aktive_df.index:
                zwischensumme = df.loc[idx, 'Zwischensumme']
                
                if zwischensumme < self.mindestbetrag:
                    # Prüfe ob Wert aus erster Runde verwendet werden soll
                    if runde > 1 and df.loc[idx, 'Erste_Berechnung'] >= self.mindestbetrag:
                        df.loc[idx, 'Endbetrag'] = df.loc[idx, 'Erste_Berechnung']
                        df.loc[idx, 'Status'] = f'Wert aus Runde 1 ({df.loc[idx, "Erste_Berechnung"]:,.2f} €)'
                        print(f"  {df.loc[idx, 'Name']}: Verwendet Wert aus Runde 1")
                    else:
                        df.loc[idx, 'Endbetrag'] = self.mindestbetrag
                        df.loc[idx, 'Status'] = f'Fixiert auf Mindestbetrag (Runde {runde})'
                        print(f"  {df.loc[idx, 'Name']}: Auf Mindestbetrag fixiert")
                    
                    neue_fixierungen.append(df.loc[idx, 'Name'])
                else:
                    df.loc[idx, 'Endbetrag'] = zwischensumme
            
            # Füge neue Fixierungen hinzu
            fixierte_kommunen.update(neue_fixierungen)
            
            # Prüfe ob Iteration beendet werden kann
            if len(neue_fixierungen) == 0:
                # Setze finale Status für nicht-fixierte Kommunen
                for idx in aktive_df.index:
                    if df.loc[idx, 'Name'] not in fixierte_kommunen:
                        df.loc[idx, 'Status'] = 'OK'
                print(f"\nBerechnung abgeschlossen nach {runde} Runden.")
                break
            
            print(f"Neu fixierte Kommunen: {len(neue_fixierungen)}")
        
        # Runde alle Endbeträge auf ganze Euro
        print("\n=== RUNDUNG AUF GANZE EURO ===")
        df['Endbetrag_vor_Rundung'] = df['Endbetrag'].copy()
        df['Endbetrag'] = df['Endbetrag'].round(0)
        
        # Berechne Differenz durch Rundung
        summe_vor_rundung = df['Endbetrag_vor_Rundung'].sum()
        summe_nach_rundung = df['Endbetrag'].sum()
        rundungs_differenz = self.gesamtsumme - summe_nach_rundung
        
        print(f"Summe vor Rundung: {summe_vor_rundung:,.2f} €")
        print(f"Summe nach Rundung: {summe_nach_rundung:,.0f} €")
        print(f"Rundungsdifferenz: {rundungs_differenz:,.0f} €")
        
        # Gleiche Differenz bei Kommune mit größter Fördersumme aus
        if abs(rundungs_differenz) >= 1:
            # Finde Kommune mit größter Fördersumme
            max_idx = df['Endbetrag'].idxmax()
            kommune_name = df.loc[max_idx, 'Name']
            alter_betrag = df.loc[max_idx, 'Endbetrag']
            
            # Korrigiere den Betrag
            df.loc[max_idx, 'Endbetrag'] += rundungs_differenz
            neuer_betrag = df.loc[max_idx, 'Endbetrag']
            
            print(f"Ausgleich bei {kommune_name}: {alter_betrag:,.0f} € → {neuer_betrag:,.0f} € (Differenz: {rundungs_differenz:+,.0f} €)")
            
            # Aktualisiere Status
            if 'Rundungsausgleich' not in df.loc[max_idx, 'Status']:
                df.loc[max_idx, 'Status'] += ' + Rundungsausgleich'
        
        # Speichere Ergebnisse
        self.ergebnis_df = df
        
        # Validierung
        self.validiere_ergebnis()
        
        return df
    
    def validiere_ergebnis(self):
        """
        Überprüft die Korrektheit der Berechnung
        """
        gesamt_verteilt = self.ergebnis_df['Endbetrag'].sum()
        differenz = abs(self.gesamtsumme - gesamt_verteilt)
        
        print(f"\n=== VALIDIERUNG ===")
        print(f"Gesamtsumme (Soll): {self.gesamtsumme:,.2f} €")
        print(f"Gesamtsumme (Ist):  {gesamt_verteilt:,.2f} €")
        print(f"Differenz:          {differenz:,.2f} €")
        
        if differenz > 0.01:  # Toleranz für Rundungsfehler
            print("⚠️ WARNUNG: Differenz größer als 1 Cent!")
        else:
            print("✓ Validierung erfolgreich!")
    
    def exportiere_excel(self, dateiname='foerdermittel_verteilung.xlsx'):
        """
        Exportiert die Ergebnisse in eine formatierte Excel-Datei
        """
        # Erstelle Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fördermittelverteilung"
        
        # Styles definieren
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titel
        ws['A1'] = "FÖRDERMITTELVERTEILUNG - BERECHNUNGSERGEBNIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:J1')
        
        # Parameter
        ws['A3'] = "Parameter:"
        ws['A3'].font = Font(bold=True)
        ws['B4'] = "Gesamtsumme:"
        ws['C4'] = self.gesamtsumme
        ws['C4'].number_format = '#,##0.00 €'
        ws['B5'] = "Mindestbetrag:"
        ws['C5'] = self.mindestbetrag
        ws['C5'].number_format = '#,##0.00 €'
        ws['B6'] = "Sockelbetrag:"
        ws['C6'] = f"{self.sockelbetrag_prozent*100}%"
        
        # Haupttabelle
        start_row = 9
        
        # Headers
        headers = ['Kommune', 'Wert 2019', 'Kinder U3', 'Sockelbetrag', 
                  'U3-Anteil', 'Zwischensumme', 'Endbetrag (gerundet)', 'Endbetrag vor Rundung', 'Status', 
                  'Runde', 'Erste Berechnung']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Daten
        for row_idx, row_data in enumerate(self.ergebnis_df.itertuples(index=False), start_row + 1):
            ws.cell(row=row_idx, column=1, value=row_data.Name).border = border
            ws.cell(row=row_idx, column=2, value=row_data.Wert_2019).border = border
            ws.cell(row=row_idx, column=2).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=3, value=row_data.Kinder_U3).border = border
            ws.cell(row=row_idx, column=4, value=row_data.Sockelbetrag).border = border
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=5, value=row_data.U3_Anteil).border = border
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=6, value=row_data.Zwischensumme).border = border
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=7, value=row_data.Endbetrag).border = border
            ws.cell(row=row_idx, column=7).number_format = '#,##0 €'  # Ganze Euro ohne Dezimalstellen
            ws.cell(row=row_idx, column=7).font = Font(bold=True)
            ws.cell(row=row_idx, column=8, value=row_data.Endbetrag_vor_Rundung).border = border
            ws.cell(row=row_idx, column=8).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=9, value=row_data.Status).border = border
            ws.cell(row=row_idx, column=10, value=row_data.Runde).border = border
            ws.cell(row=row_idx, column=11, value=row_data.Erste_Berechnung).border = border
            ws.cell(row=row_idx, column=11).number_format = '#,##0.00 €'
        
        # Summenzeile
        sum_row = start_row + len(self.ergebnis_df) + 1
        ws.cell(row=sum_row, column=1, value="SUMME").font = Font(bold=True)
        ws.cell(row=sum_row, column=7, value=self.ergebnis_df['Endbetrag'].sum())
        ws.cell(row=sum_row, column=7).number_format = '#,##0 €'  # Ganze Euro ohne Dezimalstellen
        ws.cell(row=sum_row, column=7).font = Font(bold=True)
        ws.cell(row=sum_row, column=7).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.cell(row=sum_row, column=8, value=self.ergebnis_df['Endbetrag_vor_Rundung'].sum())
        ws.cell(row=sum_row, column=8).number_format = '#,##0.00 €'
        ws.cell(row=sum_row, column=8).font = Font(bold=True)
        ws.cell(row=sum_row, column=8).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Spaltenbreiten anpassen
        column_widths = [20, 15, 12, 15, 15, 15, 18, 20, 30, 10, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Speichern
        wb.save(dateiname)
        print(f"\n✓ Excel-Datei '{dateiname}' wurde erfolgreich erstellt!")


# ========== BEISPIEL-VERWENDUNG ==========

def beispiel_berechnung():
    """
    Demonstriert die Verwendung des Fördermittelrechners mit Beispieldaten
    """
    print("=" * 60)
    print("FÖRDERMITTEL-VERTEILUNGSRECHNER")
    print("=" * 60)
    
    # Initialisiere Rechner mit Gesamtsumme
    gesamtsumme = 500000  # Beispiel: 500.000 €
    rechner = FoerdermittelRechner(gesamtsumme)
    
    # Füge Beispiel-Kommunen hinzu
    # Format: Name, Wert 2019, Anzahl Kinder U3
    beispiel_kommunen = [
        ("Stadt A", 50000, 120),
        ("Stadt B", 35000, 85),
        ("Gemeinde C", 15000, 30),
        ("Gemeinde D", 8000, 15),
        ("Stadt E", 45000, 95),
        ("Gemeinde F", 12000, 25),
        ("Stadt G", 60000, 140),
        ("Gemeinde H", 5000, 8),
        ("Stadt I", 40000, 75),
        ("Gemeinde J", 18000, 35),
    ]
    
    for name, wert_2019, kinder_u3 in beispiel_kommunen:
        rechner.kommune_hinzufuegen(name, wert_2019, kinder_u3)
    
    # Führe Berechnung durch
    print(f"\nGesamtsumme zur Verteilung: {gesamtsumme:,.2f} €")
    print(f"Anzahl Kommunen: {len(beispiel_kommunen)}")
    print("-" * 60)
    
    ergebnis = rechner.berechne_verteilung()
    
    # Zeige Zusammenfassung
    print("\n" + "=" * 60)
    print("ERGEBNIS-ÜBERSICHT")
    print("=" * 60)
    
    # Sortiere nach Endbetrag für bessere Übersicht
    ergebnis_sortiert = ergebnis.sort_values('Endbetrag', ascending=False)
    
    for _, row in ergebnis_sortiert.iterrows():
        print(f"{row['Name']:<15} → {row['Endbetrag']:>12,.0f} € | {row['Status']}")
    
    # Exportiere nach Excel
    rechner.exportiere_excel('foerdermittel_beispiel.xlsx')
    
    return rechner


# ========== FLEXIBLE EINGABE-FUNKTION ==========

def importiere_excel(dateiname):
    """
    Importiert Kommunendaten aus einer Excel-Datei
    """
    try:
        # Lese Excel-Datei ein
        df = pd.read_excel(dateiname, sheet_name="Kommunendaten", skiprows=7)
        
        # Überprüfe Spalten
        erwartete_spalten = ['Kommune', 'Wert 2019 (€)', 'Kinder U3 im SGB-II-Bezug']
        if not all(spalte in df.columns for spalte in erwartete_spalten):
            print("⚠️ Die Excel-Datei hat nicht das erwartete Format. Bitte verwenden Sie die Import-Vorlage.")
            return None
        
        # Bereinige Daten
        df = df.dropna(subset=['Kommune'])  # Entferne Zeilen ohne Kommunennamen
        df = df.rename(columns={
            'Kommune': 'Name',
            'Wert 2019 (€)': 'Wert_2019',
            'Kinder U3 im SGB-II-Bezug': 'Kinder_U3'
        })
        
        # Konvertiere Datentypen
        df['Wert_2019'] = pd.to_numeric(df['Wert_2019'], errors='coerce')
        df['Kinder_U3'] = pd.to_numeric(df['Kinder_U3'], errors='coerce')
        
        # Entferne ungültige Einträge
        df = df.dropna(subset=['Wert_2019', 'Kinder_U3'])
        
        # Überprüfe auf negative Werte
        df = df[(df['Wert_2019'] >= 0) & (df['Kinder_U3'] >= 0)]
        
        if len(df) == 0:
            print("⚠️ Keine gültigen Kommunendaten in der Excel-Datei gefunden.")
            return None
        
        print(f"✓ {len(df)} Kommunen erfolgreich aus Excel importiert.")
        return df
        
    except Exception as e:
        print(f"⚠️ Fehler beim Import der Excel-Datei: {str(e)}")
        return None

def interaktive_eingabe():
    """
    Ermöglicht die interaktive Eingabe von Kommunendaten
    """
    print("=" * 60)
    print("FÖRDERMITTEL-VERTEILUNGSRECHNER - Interaktive Eingabe")
    print("=" * 60)
    
    # Gesamtsumme eingeben
    while True:
        try:
            gesamtsumme = float(input("\nGesamtsumme zur Verteilung (in €): "))
            if gesamtsumme > 0:
                break
            else:
                print("Bitte eine positive Zahl eingeben.")
        except ValueError:
            print("Ungültige Eingabe. Bitte eine Zahl eingeben.")
    
    rechner = FoerdermittelRechner(gesamtsumme)
    
    # Eingabemethode wählen
    print("\nWie möchten Sie die Kommunendaten eingeben?")
    print("1. Manuelle Eingabe")
    print("2. Import aus Excel-Datei")
    
    while True:
        eingabe_methode = input("\nIhre Wahl (1-2): ").strip()
        
        if eingabe_methode == "1":
            # Manuelle Eingabe
            print("\nKommunendaten eingeben (leer lassen zum Beenden):")
            print("-" * 40)
            
            kommune_nr = 1
            while True:
                print(f"\nKommune {kommune_nr}:")
                name = input("  Name: ").strip()
                
                if not name:
                    if kommune_nr == 1:
                        print("Mindestens eine Kommune muss eingegeben werden.")
                        continue
                    else:
                        break
                
                try:
                    wert_2019 = float(input("  Wert 2019 (€): "))
                    kinder_u3 = float(input("  Anzahl Kinder U3 im SGB-II-Bezug: "))
                    
                    if wert_2019 < 0 or kinder_u3 < 0:
                        print("  ⚠️ Negative Werte sind nicht erlaubt. Kommune nicht hinzugefügt.")
                        continue
                        
                    rechner.kommune_hinzufuegen(name, wert_2019, kinder_u3)
                    print(f"  ✓ {name} wurde hinzugefügt.")
                    kommune_nr += 1
                    
                except ValueError:
                    print("  ⚠️ Ungültige Eingabe. Kommune nicht hinzugefügt.")
            break
            
        elif eingabe_methode == "2":
            # Excel-Import
            excel_datei = input("\nPfad zur Excel-Datei: ").strip()
            if not excel_datei:
                print("Keine Datei angegeben. Bitte versuchen Sie es erneut.")
                continue
                
            if not excel_datei.endswith('.xlsx'):
                excel_datei += '.xlsx'
                
            kommunen_df = importiere_excel(excel_datei)
            if kommunen_df is None:
                print("Import fehlgeschlagen. Bitte versuchen Sie es erneut.")
                continue
                
            # Füge Kommunen zum Rechner hinzu
            for _, row in kommunen_df.iterrows():
                rechner.kommune_hinzufuegen(row['Name'], row['Wert_2019'], row['Kinder_U3'])
            break
            
        else:
            print("Ungültige Eingabe. Bitte wählen Sie 1 oder 2.")
    
    # Berechnung durchführen
    print("\n" + "=" * 60)
    print("BERECHNUNG WIRD DURCHGEFÜHRT...")
    print("=" * 60)
    
    ergebnis = rechner.berechne_verteilung()
    
    # Ergebnis anzeigen
    print("\n" + "=" * 60)
    print("ENDERGEBNIS")
    print("=" * 60)
    
    for _, row in ergebnis.iterrows():
        print(f"{row['Name']:<20} → {row['Endbetrag']:>12,.0f} € | {row['Status']}")
    
    # Excel export
    dateiname = input("\nDateiname für Excel-Export (Standard: 'foerdermittel_verteilung.xlsx'): ").strip()
    if not dateiname:
        dateiname = 'foerdermittel_verteilung.xlsx'
    if not dateiname.endswith('.xlsx'):
        dateiname += '.xlsx'
    
    rechner.exportiere_excel(dateiname)
    
    return rechner


# ========== IMPORT-VORLAGE FUNKTION ==========

def erstelle_import_vorlage(dateiname='import_vorlage.xlsx'):
    """
    Erstellt eine Excel-Vorlage für den Import von Kommunendaten
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Kommunendaten"
    
    # Styles definieren
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titel
    ws['A1'] = "FÖRDERMITTELVERTEILUNG - IMPORT-VORLAGE"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')
    
    # Anleitung
    ws['A3'] = "Anleitung:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "1. Tragen Sie die Daten der Kommunen in die Tabelle ein."
    ws['A5'] = "2. Speichern Sie die Datei und importieren Sie sie im Programm."
    
    # Headers
    headers = ['Kommune', 'Wert 2019 (€)', 'Kinder U3 im SGB-II-Bezug']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Beispielzeile
    ws['A9'] = "Stadt Beispiel"
    ws['B9'] = 50000
    ws['C9'] = 100
    
    # Leere Zeilen für Eingabe
    for row in range(10, 30):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    
    # Speichern
    wb.save(dateiname)
    return True

# ========== HAUPTPROGRAMM ==========

if __name__ == "__main__":
    print("\nWillkommen zum Fördermittel-Verteilungsrechner!")
    print("\nWählen Sie eine Option:")
    print("1. Beispielberechnung mit Testdaten durchführen")
    print("2. Eigene Daten eingeben (manuell oder Excel-Import)")
    print("3. Import-Vorlage für Excel erstellen")
    print("4. Programm beenden")
    
    while True:
        wahl = input("\nIhre Wahl (1-4): ").strip()
        
        if wahl == "1":
            beispiel_berechnung()
            break
        elif wahl == "2":
            ergebnis = interaktive_eingabe()
            if ergebnis is None:
                print("\nProgramm wurde ohne Berechnung beendet.")
            break
        elif wahl == "3":
            vorlage_name = input("\nDateiname für Import-Vorlage (Standard: 'import_vorlage.xlsx'): ").strip()
            if not vorlage_name:
                vorlage_name = 'import_vorlage.xlsx'
            if not vorlage_name.endswith('.xlsx'):
                vorlage_name += '.xlsx'
            erstelle_import_vorlage(vorlage_name)
            print(f"\n✓ Die Vorlage wurde erstellt: {vorlage_name}")
            print("  Füllen Sie die Vorlage aus und wählen Sie dann Option 2 für den Import.")
            
            if input("\nMöchten Sie jetzt mit dem Import fortfahren? (j/n): ").lower() == 'j':
                interaktive_eingabe()
            break
        elif wahl == "4":
            print("Programm wird beendet.")
            break
        else:
            print("Ungültige Eingabe. Bitte 1, 2, 3 oder 4 wählen.")
    
    print("\nVielen Dank für die Nutzung des Fördermittel-Verteilungsrechners!")